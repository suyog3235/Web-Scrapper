import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup


def facebook():

    # facebook login data
    USERNAME = 'Your facebook username'
    PASSWORD = 'your facebook password'

    # turn off all the facebook popup notifications
    chrome_options = webdriver.ChromeOptions()
    prefs = {"profile.default_content_setting_values.notifications": 2}
    chrome_options.add_experimental_option("prefs", prefs)

    # FOR EXAMPLE  driver = webdriver.Chrome(options=chrome_options, executable_path=r"C:\Users\XD\Desktop\internship program\chromedriver.exe")
    driver = webdriver.Chrome(options=chrome_options,
                              executable_path=r"ADD CHROME DRIVER PATH HERE SEE EXAMPLE ABOVE")
    driver.maximize_window()

    # login to face book (using m basic here because its easier to scrape with mbasic )
    driver.get("https://mbasic.facebook.com/login/")
    time.sleep(2)
    driver.find_element_by_id("m_login_email").send_keys(USERNAME)
    time.sleep(1)
    driver.find_element_by_name("pass").send_keys(PASSWORD)
    time.sleep(2)
    driver.find_element_by_name("login").click()
    time.sleep(5)
    driver.get("https://mbasic.facebook.com/CloudCounselage")
    time.sleep(2)

    # scroll down to load the whole page
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:

        # Scroll down to the bottom.
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);")

        # Wait to load the page.
        time.sleep(5)

        # Calculate new scroll height and compare with last scroll height.
        new_height = driver.execute_script("return document.body.scrollHeight")

        if new_height == last_height:
            break

        last_height = new_height

    src = driver.page_source
    soup = BeautifulSoup(src, 'lxml')

    total_fb_posts_in_july = 0
    total_fb_likes_in_july = 0
    total_fb_comments_in_july = 0

    try:
        posts = soup.find_all("article", attrs={"class": "db ej gl"})
        for post in posts:
            d = post.find_all("abbr")
            for d1 in d:
                if "July" in d1.text:
                    total_fb_posts_in_july += 1
                    likes = post.find_all("a", attrs={"class": "hf hg"})
                    for like in likes:
                        total_fb_likes_in_july += int(like.text)

                    comments = post.find_all("a", attrs={"class": "hg"})
                    for comment in comments:
                        if "Comment" in comment.text:
                            total_fb_comments_in_july += 1
    except Exception as e:
        print("error while getting data from facebook !!!!")
        print("Error:", e)

    print("success !!!")
    wb = load_workbook('Audit Report For DM.xlsx')

    sheet = wb.active
    sheet['C2'] = total_fb_posts_in_july
    sheet['D2'] = total_fb_likes_in_july
    sheet['E2'] = total_fb_comments_in_july
    wb.save('Audit Report For DM.xlsx')
    print("saved facebook data into excel file.")

    # log out and close the window
    time.sleep(5)
    driver.find_element_by_id("mbasic_logout_button").click()
    time.sleep(5)
    driver.close()


def linkedin():

    # FOR EXAMPLE:  driver = webdriver.Chrome(r"C:\Users\XD\Desktop\internship program\chromedriver.exe")
    driver = webdriver.Chrome(
        r"ADD CHROME DRIVER PATH HERE SEE EXAMPLE ABOVE")
    driver.maximize_window()

    # linkdin account info
    USERNAME = 'your linkdin username'
    PASSWORD = 'your linkdin password'

    driver.get("https://www.linkedin.com/login")
    # driver.refresh()
    time.sleep(2)
    email = driver.find_element_by_id("username")
    email.send_keys(USERNAME)
    password = driver.find_element_by_id("password")
    password.send_keys(PASSWORD)
    time.sleep(2)
    password.send_keys(Keys.RETURN)
    time.sleep(2)

    driver.get(
        "https://www.linkedin.com/company/cloud-counselage/posts/?feedView=all")
    time.sleep(5)

    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:

        # Scroll down to the bottom.
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);")

        # Wait to load the page.
        time.sleep(2)

        # Calculate new scroll height and compare with last scroll height.
        new_height = driver.execute_script("return document.body.scrollHeight")

        if new_height == last_height:
            break

        last_height = new_height

    time.sleep(1)
    src = driver.page_source
    soup = BeautifulSoup(src, "lxml")

    total_posts = soup.find_all("div", attrs={"class": "occludable-update"})
    total_posts_in_july = 0
    total_likes = 0
    total_comments = 0

    try:
        for post in total_posts:
            day_time = post.find_all(
                "span", attrs={"class": "visually-hidden"})
            for d in day_time:
                if d.text == "2 months ago":
                    total_posts_in_july += 1

                    likes = post.find_all("span", attrs={
                                          'class': "v-align-middle social-details-social-counts__reactions-count"})
                    for like in likes:
                        total_likes += int(like.text)

                    comments = post.find_all(
                        "", attrs={"class": "social-details-social-counts__comments"})
                    for comment in comments:
                        total_comments += int(comment.text)
    except Exception as e:
        print("failed to get data from linkdin")
        print("Error:", e)

    print("success !!!")
    wb = load_workbook('Audit Report For DM.xlsx')

    sheet = wb.active
    sheet['C3'] = total_posts_in_july
    sheet['D3'] = total_likes
    sheet['E3'] = total_comments
    wb.save('Audit Report For DM.xlsx')
    print("saved linkdin data into excel file.")

    driver.close()


facebook()
linkedin()
